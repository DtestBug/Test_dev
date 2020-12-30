from django.http.response import HttpResponse
from rest_framework.views import APIView
from utils.get_excel_stream import get_excel_stream
from utils.read_excel import read_excel

from openpyxl import load_workbook
from openpyxl.styles import PatternFill  # 导入填充模块
import configparser
import requests
from urllib import parse


class FileUpload(APIView):

    def __init__(self):
        self.web1 = configparser.ConfigParser()
        self.web1.read("data_detection\data_set.conf", encoding='utf-8')
        self.test_url = self.web1.get("set", "API_url")

    def post(self, request):
        paper_file = request.FILES['file']  # excel文件绝对路径

        rewrite = load_workbook(paper_file)  # 读取文件
        sheets = rewrite.sheetnames  # 获取所有sheet
        res = {}
        count = 0
        for sheet_name in sheets:
            count += 1
            res[f'sheet{count}'] = sheet_name  # 汇总所有sheet表单
            # ========================================================
            print('=' * 100)
            print(f'\033[1;34m sheet_name: {sheet_name} \033[0m')
            print('=' * 100)

            # 追加数据需要调用已读取的 excel路径 >>> sheet_name
            ws1 = rewrite[sheet_name]

            # 先写头部
            colour_yellow = PatternFill('solid', fgColor='99FF00')  # darkVertical 是填充竖线  99FF00是黄色
            excel_header = ['接口原价', '接口售价', '原价对比结果', '售价对比结果']  # 首行头部信息作为变量可更改

            # 从第XX列开始横向插入title
            header_count = int(self.web1.get("set", "header_count"))
            for sheet_col in range(len(excel_header)):
                header_count += 1
                ws1.cell(row=1, column=header_count).value = excel_header[sheet_col]  # 依次增加首行数据
                ws1.cell(row=1, column=header_count).fill = colour_yellow  # 首行颜色

            m = 0
            file_data = read_excel(paper_file, sheet_name)
            for data in file_data:
                # =============以上内容通用================
                # 接口请求内容
                sheet_id = data['活动id']
                if sheet_id is None:
                    m += 1
                    ws1.cell(row=m + 1, column=int(self.web1.get("set", "column_1"))).value = ' '  # excel追加数据
                    ws1.cell(row=m + 1, column=int(self.web1.get("set", "column_2"))).value = ' '  # excel追加数据
                    ws1.cell(row=m + 1, column=int(self.web1.get("set", "column_3"))).value = ' '  # excel追加数据
                    ws1.cell(row=m + 1, column=int(self.web1.get("set", "column_4"))).value = ' '  # excel追加数据
                else:
                    payload = "{\"campaignId\":\"%s\"}" % sheet_id
                    headers = {
                        'Content-Type': 'application/json',
                        'cookie': 'BIDUPSID=F3483E1FDBB901C21F470813431C63C8; BDORZ=B490B5EBF6F3CD402E515D22BCDA1598; BAIDUID=174A208A202865FD3F1D83FD553758C7:FG=1; BAIDUID_BFESS=174A208A202865FD3F1D83FD553758C7:FG=1; PSTM=1608700874; delPer=0; H_PS_PSSID=1446_33356_33306_31660_33351_33313_33312_33311_33310_33309_26350_33308_33307_22159_33389; PSINO=2; ab_sr=1.0.0_NTQ3N2JlZjNmNTA4ZWUyNDA5NTE0Yzk5ZjcyOGVjYTBkNTY0NTQ3OTZlZWE2Yzk1MTIzMGJmY2E5M2Q5ZTVkM2YzNDAzNmU3NDMxN2IyNjE0MTY2MTYwYTc2OGFkYjQ3; __yjsv5_shitong=1.0_7_9cd687f645a15514793fad34031c1f811490_300_1608707110609_111.206.214.32_322498fa; yjs_js_security_passport=a336b10b371ac3ad40c1958e862521b45118f277_1608707111_js; __xsptplusUT_861=1; UUAP_P_TOKEN=PT-548530374738202624-SawisMzfDu-uuap; SECURE_UUAP_P_TOKEN=PT-548530374738202624-SawisMzfDu-uuap; BSG_B_TOKEN=3UaasTcl7sy9jCzBPHgOzhbX/N5acErg8s4mP3RCnlP7Sp/tKi7oU7CC5LKr6garPeHqDpvmRAaGPmjeiwavIscy2A0VpcY1q5UhYZ3SXow=; SECURE_BSG_B_TOKEN=3UaasTcl7sy9jCzBPHgOzhbX/N5acErg8s4mP3RCnlP7Sp/tKi7oU7CC5LKr6garPeHqDpvmRAaGPmjeiwavIscy2A0VpcY1q5UhYZ3SXow=; BIDUPSID_BFESS=F3483E1FDBB901C21F470813431C63C8; AGL_USER_ID=80188b14-75e2-44a5-b220-ff1cb9210c01; _ga=GA1.2.2026268364.1608709438; _gid=GA1.2.546434689.1608709438; BDUSS=UxN0Q2aHp1Vm5ETUp1a1hiMX5iZFRyWWlYeTlLTm9jZWVVQ0YtWkdNeTdnZ3BnRUFBQUFBJCQAAAAABwAAAAEAAACgxzCRAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAALv14l-79eJfYm; BDUSS_BFESS=UxN0Q2aHp1Vm5ETUp1a1hiMX5iZFRyWWlYeTlLTm9jZWVVQ0YtWkdNeTdnZ3BnRUFBQUFBJCQAAAAABwAAAAEAAACgxzCRAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAALv14l-79eJfYm; __xsptplus861=861.1.1608709424.1608709567.3%234%7C%7C%7C%7C%7C%23%23s56MYxqU4qcUG0XoDYMzbeTuiAs9bfnw%23; Hm_lvt_28a17f66627d87f1d046eae152a1c93d=1608709436,1608709567; Hm_lpvt_28a17f66627d87f1d046eae152a1c93d=1608709567; _gat_gtag_UA_138572523_1=1; bce-userbind-source=PASSPORT%3BUUAP; bce-ctl-client-cookies="BDUSS,bce-passport-stoken,bce-device-cuid,bce-device-token,BAIDUID"; bce-auth-type=PASSPORT; bce-passport-stoken=75fd54f6c390a0d356c86699e5c78bfdfe4e6fa51a068848c5388acdfed0880f; bce-sessionid=001c21cc4b36d144023bc6f6e6b5e61d4d2; bce-user-info=2020-12-23T15:46:05Z|d50787b7ce807d925a3e38ce9228d225; bce-ctl-sessionmfa-cookie=bce-session; bce-session=fffc58b7eabd4daa8937269dc27f62bf209e6950d170471086e3872368d639c2|30e5fbf68abb5c7e6c680425eb158e23; bce-login-type=PASSPORT; bce-login-expire-time="2020-12-23T08:16:06Z|c71b5d372d5190a4d5559fb980350264"; bce-login-display-name=*******5733; BAIDU_CLOUD_TRACK_PATH=https%3A%2F%2Fcloudtest.baidu.com%2Fcampaign%2F20201212-ptcj%2Findex.html%3F_%3D1608709564970'}
                    response = requests.request("POST", self.test_url, data=payload, headers=headers)

                    if response.status_code == 500:
                        print(f'\033[1;33m接口返回,code: {response.status_code}\033[0m  Excel_id:{payload}', )
                        m += 1
                        ws1.cell(row=m + 1, column=int(self.web1.get("set", "column_1"))).value = '返回500'  # excel追加数据
                        ws1.cell(row=m + 1, column=int(self.web1.get("set", "column_2"))).value = '返回500'  # excel追加数据

                    if response.status_code == 200:
                        print(payload, response.json())

                        # 返回结果为200，并且有result结果
                        if 'result' in response.json():
                            if data['活动价'] is None or data['目录价'] is None:
                                ws1.cell(row=m + 1,
                                         column=int(self.web1.get("set", "column_1"))).value = ' '  # excel追加数据
                                ws1.cell(row=m + 1,
                                         column=int(self.web1.get("set", "column_2"))).value = ' '  # excel追加数据
                            else:
                                # excel价格信息
                                sheet_old_price = float(data['目录价'])
                                sheet_price = float(data['活动价'])

                                # 接口价格信息
                                api_old_price = float(response.json()['result']['originalPrice'])
                                api_price = float(response.json()['result']['campaignPrice'])

                                m += 1
                                # 价格存入excel
                                ws1.cell(row=m + 1,
                                         column=int(self.web1.get("set", "column_1"))).value = api_old_price  # excel追加数据
                                ws1.cell(row=m + 1,
                                         column=int(self.web1.get("set", "column_2"))).value = api_price  # excel追加数据
                                # 存入判断结果
                                if api_old_price != sheet_old_price:
                                    differ = api_old_price - sheet_old_price  # 计算差价
                                    ws1.cell(row=m + 1, column=int(
                                        self.web1.get("set", "column_3"))).value = '原价不一致，相差 %s' % differ  # excel追加数据
                                else:
                                    ws1.cell(row=m + 1,
                                             column=int(self.web1.get("set", "column_3"))).value = 'success'  # excel追加数据

                                if api_price != sheet_price:
                                    differ = api_price - sheet_price  # 计算差价
                                    ws1.cell(row=m + 1, column=int(
                                        self.web1.get("set", "column_4"))).value = '售价不一致，相差 %s' % differ  # excel追加数据
                                else:
                                    ws1.cell(row=m + 1,
                                             column=int(self.web1.get("set", "column_4"))).value = 'success'  # excel追加数据

                        # 返回结果为200，没有有result结果
                        else:
                            m += 1
                            ws1.cell(row=m + 1, column=int(self.web1.get("set", "column_1"))).value = '无结果返回code:%s' % \
                                                                                                 response.json()[
                                                                                                     'status']  # excel追加数据
                            ws1.cell(row=m + 1, column=int(self.web1.get("set", "column_2"))).value = '无结果返回code:%s' % \
                                                                                                 response.json()[
                                                                                                     'status']  # excel追加数据
                            print('无结果返回', payload, response.json())
        # rewrite.save(paper_file)  # 保存追加数据后的表

        res = get_excel_stream(rewrite)
        # 设置HttpResponse的类型
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment;filename=' + parse.quote("New_excel") + '.xlsx'
        # 将文件流写入到response返回
        response.write(res)

        return response


