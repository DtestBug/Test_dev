from openpyxl import load_workbook


def read_excel(file_path, sheet_name):
    # 读取Excel表格表面数据，不读取公式，如需获取表格内公式，将data_only改为false(本代码不推荐，因为改了之后不会显示Excel数据)
    file_sheet = load_workbook(file_path, data_only=True,)[sheet_name]

    # 获取头部
    sheet_header = []
    for Header_data in range(file_sheet.max_column):
        sheet_header.append(file_sheet.cell(row=1, column=Header_data+1).value)

    # 获取数据
    data_list = []
    for data_data1 in range(file_sheet.max_row-1):
        data_dict = {}
        for data_data2 in range(len(sheet_header)):
            data_dict[sheet_header[data_data2]] = file_sheet.cell(row=data_data1+2, column=data_data2+1).value
        data_list.append(data_dict)
    return data_list  # 返回数据
